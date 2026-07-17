"""One-time offline converter: data/*.xlsx -> data/listings.json.

Mechanically parses excel sheets and normalizes rows using keyword matching.
No external API calls. Run: python convert_data.py

Output shape: {"indore": [listing, ...], "mumbai": [...], "delhi": [...]}
"""
import json
import logging
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("covio2.convert")

DATA_DIR = Path(__file__).parent / "data"
OUT_PATH = DATA_DIR / "listings.json"
CITIES = {"indore": "IND", "mumbai": "MUM", "delhi": "DEL"}
EXCEL_EPOCH = datetime(1899, 12, 30)

HEADER_PATTERNS = [
    (re.compile(r"OBJECT", re.I), "object"),
    (re.compile(r"LOCATION", re.I), "location"),
    (re.compile(r"PRICE", re.I), "price"),
    (re.compile(r"DELIVERY", re.I), "delivery"),
    (re.compile(r"CONTACT|PHONE", re.I), "phone"),
    # re-verification checked before the generic DATE so the newer date wins
    (re.compile(r"RE.?VERIF", re.I), "reverified_date"),
    (re.compile(r"DATE|VERIF", re.I), "verified_date"),
    (re.compile(r"DOCUMENT|REQUIREMENT", re.I), "documents"),
]

RESOURCE_KEYWORDS = [
    ("cylinder", re.compile(r"c[yi]l[ie]nder|cyclinder", re.I)),
    ("concentrator", re.compile(r"concentrator", re.I)),
    ("refill", re.compile(r"refill", re.I)),
    ("can", re.compile(r"\bcan\b", re.I)),
]

DELIVERY_PATTERNS = [
    ("both", re.compile(r"both|home.*pick|pick.*home", re.I)),
    ("pickup", re.compile(r"pick\s*up|no\s*home\s*deliver|self\s*collect", re.I)),
    ("home_delivery", re.compile(r"home\s*deliver|deliver.*home|door\s*step", re.I)),
]

INACTIVE_PATTERNS = [
    ("dead or unreachable phone number",
     re.compile(r"doesn.?t\s*exist|not\s*exist|switched\s*off|unreachable|wrong\s*number", re.I)),
    ("stock not available",
     re.compile(r"none\s*available|not\s*available|out\s*of\s*stock", re.I)),
    ("hospital-internal supply only",
     re.compile(r"hospital\s*(internal|use)\s*only|for\s*patients\s*only", re.I)),
]

DOCS_NONE_PATTERNS = re.compile(
    r"no\s*require|none\s*needed|no\s*such\s*verif|no\s*document|not\s*required|nil|none",
    re.I,
)


def parse_phone(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(int(float(value)))
    text = str(value).strip()
    if not text:
        return None
    try:
        return str(int(float(text)))
    except ValueError:
        digits = re.sub(r"\D", "", text)
        return digits if len(digits) >= 7 else None


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)):
        try:
            return (EXCEL_EPOCH + timedelta(days=float(value))).date().isoformat()
        except (ValueError, OverflowError):
            return None
    text = str(value).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def normalize_delivery(text):
    if not text:
        return "unknown"
    for label, pattern in DELIVERY_PATTERNS:
        if pattern.search(text):
            return label
    return "unknown"


def normalize_documents(text):
    if not text:
        return None
    if DOCS_NONE_PATTERNS.search(text):
        return "none"
    return text.strip()


def detect_inactive(raw):
    if not raw["phone"]:
        return "missing phone number"
    obj_text = (raw["object"] or "") + " " + (raw["price"] or "")
    for reason, pat in INACTIVE_PATTERNS:
        if pat.search(obj_text):
            return reason
    return None


def normalize_object_text(text):
    if not text:
        return text
    text = re.sub(r"\s+", " ", text.strip())
    text = text[0].upper() + text[1:] if text else text
    common_fixes = [
        (re.compile(r"oxgy?gen|oxyg?en|oxgen", re.I), "Oxygen"),
        (re.compile(r"c[yi]l[ie]nder|cyclinder", re.I), "Cylinder"),
        (re.compile(r"concentrater", re.I), "Concentrator"),
    ]
    for pat, replacement in common_fixes:
        text = pat.sub(replacement, text)
    return text


def parse_sheet(path, city):
    from openpyxl import load_workbook

    log.info("parsing %s", path.name)
    ws = load_workbook(path, read_only=True, data_only=True).worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    log.info("  %s: %d raw rows in first worksheet", path.name, len(rows))
    header_idx, col_map = None, {}
    for i, row in enumerate(rows):
        cells = [str(c) for c in row if c is not None]
        if any(re.search(r"OBJECT", c, re.I) for c in cells):
            header_idx = i
            for j, cell in enumerate(row):
                if cell is None:
                    continue
                for pattern, field in HEADER_PATTERNS:
                    if pattern.search(str(cell)) and field not in col_map:
                        col_map[field] = j
            break
    if header_idx is None:
        log.warning("no header row (cell matching /OBJECT/i) found in %s, skipping", path.name)
        return []
    log.info("  header found at row %d, columns mapped: %s", header_idx + 1, col_map)
    missing = [f for _, f in HEADER_PATTERNS if f not in col_map]
    if missing:
        log.warning("  %s: no column matched for %s — those fields will be null", path.name, missing)

    parsed = []
    for row in rows[header_idx + 1:]:
        if sum(1 for c in row if c is not None and str(c).strip()) < 3:
            continue
        get = lambda f: row[col_map[f]] if f in col_map and col_map[f] < len(row) else None
        # prefer the re-verification date (newer) when the sheet has one
        verified = parse_date(get("reverified_date")) or parse_date(get("verified_date"))
        parsed.append({
            "id": f"{CITIES[city]}-{len(parsed) + 1:03d}",
            "city": city,
            "object": str(get("object") or "").strip() or None,
            "location": str(get("location") or "").strip() or None,
            "price": str(get("price") or "").strip() or None,
            "delivery": str(get("delivery") or "").strip() or None,
            "phone": parse_phone(get("phone")),
            "verified_date": verified,
            "documents": str(get("documents") or "").strip() or None,
        })
    return parsed


def build_listing(raw):
    types = [t for t, pat in RESOURCE_KEYWORDS if raw["object"] and pat.search(raw["object"])]
    inactive_reason = detect_inactive(raw)
    return {
        "id": raw["id"],
        "city": raw["city"],
        "resource_types": types or ["other"],
        "raw_object": normalize_object_text(raw["object"]),
        "location": raw["location"],
        "price": raw["price"],
        "is_free": bool(raw["price"] and re.search(r"free", raw["price"], re.I)),
        "delivery": normalize_delivery(raw["delivery"]),
        "phone": raw["phone"],
        "verified_date": raw["verified_date"],
        "documents": normalize_documents(raw["documents"]),
        "active": inactive_reason is None,
        "inactive_reason": inactive_reason,
    }


def main():
    log.info("=== convert_data start ===")
    listings, summary = {}, []
    for city in CITIES:
        path = DATA_DIR / f"{city}.xlsx"
        if not path.exists():
            log.info("%s not found, skipping %s", path.name, city)
            continue
        raw_rows = parse_sheet(path, city)
        log.info("  %s: %d usable rows after skipping blanks", city, len(raw_rows))
        city_listings = [build_listing(r) for r in raw_rows]
        listings[city] = city_listings
        active = sum(1 for r in city_listings if r["active"])
        inactive = len(raw_rows) - active
        summary.append((city, len(raw_rows), active, inactive))
        log.info("  %s: %d active, %d inactive", city, active, inactive)

    if not listings:
        log.error("no xlsx files found in data/ — nothing to convert")
        sys.exit(1)

    total = sum(len(rows) for rows in listings.values())
    OUT_PATH.write_text(json.dumps(listings, indent=2, ensure_ascii=False))
    log.info("=== convert_data done ===")
    log.info("wrote %d listings across %d cities to %s", total, len(listings), OUT_PATH)
    print(f"\n{'city':<10}{'parsed':>8}{'active':>8}{'inactive':>10}")
    for city, parsed, active, inactive in summary:
        print(f"{city:<10}{parsed:>8}{active:>8}{inactive:>10}")


if __name__ == "__main__":
    main()
